#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""人类专家盲测的效度分析 —— 外审「参考集可能循环论证」这一条的正面回应。

三个产出:
  (1) 人-人 Cohen's κ  —— 两位**真正独立**的人类标注者之间, 顶替此前那个
      「两个被提示的模型实例」的 κ
  (2) 人-模型一致率, **按分层报告** —— 回答"参考集在哪些地方可信、哪些地方不可信"
  (3) 专家共识子集上的主结果 —— 把评测限制在两位专家都同意模型判定的字段上,
      看结论是否仍然成立

输出: results/tables/table5_expert_validation.csv
      results/tables/table5_expert_disagreements.csv  (逐条分歧, 供论文举例与后续修正)
"""
import collections, csv, json, os, sys
from openpyxl import load_workbook

HV = "human_validation/_filled"
OUT = "results/tables"
SRC = {"clinician": os.path.join(HV, "给临床专家/标注表_临床专家.xlsx"),
       "data_eng": os.path.join(HV, "给数据工程专家/标注表_数据工程专家.xlsx")}


def norm(x):
    t = str(x or "").strip()
    if not t:
        return ""
    u = t.upper()
    if u in ("UNKNOWN", "UNK", "NONE", "N/A"):
        return "UNKNOWN"
    if u in ("UNSURE", "UNCERTAIN", "?"):
        return "UNSURE"
    return t.strip().lower().replace(" ", "_").replace("-", "_")


def read(p):
    wb = load_workbook(p); s = wb["标注表"]
    hdr = [c.value for c in s[1]]
    i_no = hdr.index("item_no"); i_a = hdr.index("YOUR_ANSWER_concept_or_UNKNOWN")
    i_c = hdr.index("YOUR_CONFIDENCE_1to5"); i_n = hdr.index("YOUR_NOTE")
    out = {}
    for r in s.iter_rows(min_row=2):
        v = [c.value for c in r]
        out[int(v[i_no])] = {"ans": norm(v[i_a]),
                             "conf": int(v[i_c]) if str(v[i_c] or "").strip().isdigit() else None,
                             "note": str(v[i_n] or "").strip()}
    return out


def kappa(a, b):
    """Cohen's κ, 类别 = 具体概念名 / UNKNOWN。UNSURE 行按惯例剔除。"""
    pairs = [(x, y) for x, y in zip(a, b) if x not in ("", "UNSURE") and y not in ("", "UNSURE")]
    n = len(pairs)
    po = sum(1 for x, y in pairs if x == y) / n
    ca = collections.Counter(x for x, _ in pairs); cb = collections.Counter(y for _, y in pairs)
    pe = sum(ca[k] * cb.get(k, 0) for k in ca) / (n * n)
    return n, po, pe, (po - pe) / (1 - pe) if pe < 1 else float("nan")


if __name__ == "__main__":
    key = {int(r["item_no"]): r for r in csv.DictReader(
        open("human_validation/_sample_key_DO_NOT_SEND.csv", newline="", encoding="utf-8"))}
    A, B = read(SRC["clinician"]), read(SRC["data_eng"])
    items = sorted(key)
    ka = [A[i]["ans"] for i in items]
    kb = [B[i]["ans"] for i in items]
    km = [norm(key[i]["model_answer"]) for i in items]

    rows = []
    n, po, pe, k = kappa(ka, kb)
    print("=== (1) 人-人 一致性 (两位独立人类标注者) ===")
    print("    n=%d  Po=%.4f  Pe=%.4f  Cohen's κ = %.4f" % (n, po, pe, k))
    rows.append({"comparison": "clinician vs data engineer (human-human)", "stratum": "all",
                 "n": n, "Po": round(po, 4), "Pe": round(pe, 4), "kappa": round(k, 4)})
    # 二值化 (可映射 vs UNKNOWN) 的一致性 —— 开放集判定本身
    bin_a = ["M" if x not in ("UNKNOWN", "UNSURE", "") else x for x in ka]
    bin_b = ["M" if x not in ("UNKNOWN", "UNSURE", "") else x for x in kb]
    n2, po2, pe2, k2 = kappa(bin_a, bin_b)
    print("    二值(可映射 vs UNKNOWN): n=%d Po=%.4f Pe=%.4f κ=%.4f" % (n2, po2, pe2, k2))
    rows.append({"comparison": "clinician vs data engineer (mappable vs UNKNOWN)",
                 "stratum": "all", "n": n2, "Po": round(po2, 4), "Pe": round(pe2, 4),
                 "kappa": round(k2, 4)})

    print("\n=== (2) 人-模型 一致率, 按分层 ===")
    print("    %-6s %-4s %-24s %-24s %s" % ("层", "n", "临床专家 vs 模型", "数据工程 vs 模型", "两人均同意模型"))
    order = ["S1", "S2", "S3", "S4", "S5"]
    desc = {"S1": "判 UNKNOWN 的高覆盖字段", "S2": "正例·仅 LLM 证据",
            "S3": "两标注者分歧经仲裁", "S4": "检查报冲突的正例", "S5": "正例·非 LLM 证据"}
    for st in order + ["ALL"]:
        sel = [i for i in items if st == "ALL" or key[i]["stratum"] == st]
        if not sel:
            continue
        aa = sum(1 for i in sel if A[i]["ans"] == norm(key[i]["model_answer"]))
        bb = sum(1 for i in sel if B[i]["ans"] == norm(key[i]["model_answer"]))
        both = sum(1 for i in sel if A[i]["ans"] == B[i]["ans"] == norm(key[i]["model_answer"]))
        print("    %-6s %-4d %-24s %-24s %s"
              % (st, len(sel), "%5.1f%% (%d/%d)" % (100 * aa / len(sel), aa, len(sel)),
                 "%5.1f%% (%d/%d)" % (100 * bb / len(sel), bb, len(sel)),
                 "%5.1f%%" % (100 * both / len(sel))))
        rows.append({"comparison": "human vs model", "stratum": "%s (%s)" % (st, desc.get(st, "all")),
                     "n": len(sel), "clinician_agree_%": round(100 * aa / len(sel), 1),
                     "data_eng_agree_%": round(100 * bb / len(sel), 1),
                     "both_agree_%": round(100 * both / len(sel), 1)})

    # (3) 逐条分歧
    dis = []
    for i in items:
        m = norm(key[i]["model_answer"])
        if A[i]["ans"] == m and B[i]["ans"] == m:
            continue
        dis.append({"item_no": i, "stratum": key[i]["stratum"], "db": key[i]["db"],
                    "field_key": key[i]["field_key"], "model": m,
                    "clinician": A[i]["ans"], "clinician_conf": A[i]["conf"],
                    "data_eng": B[i]["ans"], "data_eng_conf": B[i]["conf"],
                    "both_humans_agree_against_model":
                        A[i]["ans"] == B[i]["ans"] and A[i]["ans"] != m,
                    "clinician_note": A[i]["note"][:160], "data_eng_note": B[i]["note"][:160]})
    strong = [d for d in dis if d["both_humans_agree_against_model"]]
    print("\n=== (3) 分歧 ===")
    print("    与模型不完全一致: %d / %d" % (len(dis), len(items)))
    print("    **两位专家一致地反对模型**: %d / %d (%.1f%%)  <- 这是模型真正可能错的地方"
          % (len(strong), len(items), 100 * len(strong) / len(items)))
    print("    分层分布:", dict(collections.Counter(d["stratum"] for d in strong)))
    d2 = collections.Counter()
    for d in strong:
        if d["model"] == "UNKNOWN":
            d2["模型判 UNKNOWN, 专家给了概念"] += 1
        elif d["clinician"] == "UNKNOWN":
            d2["模型给了概念, 专家判 UNKNOWN"] += 1
        else:
            d2["双方都给概念但不同"] += 1
    print("    方向:", dict(d2))

    os.makedirs(OUT, exist_ok=True)
    cols = ["comparison", "stratum", "n", "Po", "Pe", "kappa",
            "clinician_agree_%", "data_eng_agree_%", "both_agree_%"]
    with open(os.path.join(OUT, "table5_expert_validation.csv"), "w", newline="",
              encoding="utf-8") as f:
        w = csv.DictWriter(f, cols); w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})
    with open(os.path.join(OUT, "table5_expert_disagreements.csv"), "w", newline="",
              encoding="utf-8") as f:
        w = csv.DictWriter(f, list(dis[0].keys())); w.writeheader(); w.writerows(dis)
    print("\n-> %s/table5_expert_validation.csv (+ _disagreements.csv)" % OUT)
